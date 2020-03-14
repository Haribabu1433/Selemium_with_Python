import openpyxl

path = r"C:\Users\hpolicha\Documents\pyexcel\data1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet =workbook.active                      #workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column

print(rows,cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
            print(str(sheet.cell(row=r,column=c).value).rjust(5),end="   ")
    print()